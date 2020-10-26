import pygame
import random
from pygame.draw import *
from random import randint
import openpyxl

pygame.init()

FPS = 60
screen = pygame.display.set_mode((1200, 900))

RED = (255, 0, 0)
BLUE = (0, 0, 255)
YELLOW = (255, 255, 0)
GREEN = (0, 255, 0)
MAGENTA = (255, 0, 255)
CYAN = (0, 255, 255)
LIGHT_BLUE = (50, 204, 255)
LIGHT_GREY = (240, 240, 250)
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
COLORS = [RED, BLUE, YELLOW, GREEN, MAGENTA, CYAN, LIGHT_BLUE]
COLORS_2 = [BLUE, RED, GREEN]


class Score:

    def __init__(self):
        self.score = 0

    def increase(self, points):
        """Increases score for the needed for the aim points"""
        self.score += points

    def decrease(self):
        """Decreases score for the needed for the aim points"""
        self.score -= self.score

    def text(self):
        """Writes the current score on the screen"""
        surface_score = pygame.font.SysFont('Helvetic', 100).render(str(self.score), False, BLACK)
        screen.blit(surface_score, (50, 50))

    def hello(self):
        """Writes the rules on the screen"""
        surface_hi = pygame.font.SysFont('Helvetic', 50).render('Do not approach my car', False, BLACK)
        screen.blit(surface_hi, (700, 50))


class Ball(object):

    def __init__(self):
        self.color = random.choice(COLORS_2)
        self.x = randint(100, 1000)
        self.y = randint(100, 800)
        self.r = randint(30, 50)
        self.speed_x = randint(-100, 100)
        self.speed_y = randint(-100, 100)

    def draw_ball(self):
        """Draws a ball of a random color in random x, y coordinates with random radius"""
        circle(screen, self.color, (self.x, self.y), self.r)

    def move(self):
        """Moves the ball in the screen (with reflection from the walls)"""
        self.x += self.speed_x / FPS
        self.y += self.speed_y / FPS
        self.draw_ball()
        if self.x >= 1100:
            self.speed_x = randint(-100, -10)
        if self.x <= 50:
            self.speed_x = randint(10, 100)
        if self.y >= 800:
            self.speed_y = randint(-100, -10)
        if self.y <= 50:
            self.speed_y = randint(10, 100)

    def click(self, pos):
        """Finds out if the player caught the ball"""
        x, y = pos
        if (self.x - x) ** 2 + (self.y - y) ** 2 <= self.r ** 2:
            self.color = random.choice(COLORS)
            self.x = randint(100, 1000)
            self.y = randint(100, 800)
            self.r = randint(30, 50)
            self.speed_x = randint(-100, 100)
            self.speed_y = randint(-100, 100)
            return True
        else:
            return False


class Aim(object):

    def __init__(self):
        self.color = random.choice(COLORS)
        self.x = randint(100, 1000)
        self.y = randint(100, 800)
        self.r = randint(50, 200)
        self.speed_x = randint(-200, 200)
        self.speed_y = randint(-200, 200)

    def draw_aim(self):
        """Draws a special aim of a random color in random x, y coordinates with random size"""
        polygon(screen, self.color, [(self.x, self.y), (self.x + self.r * 1.71 / 2, self.y - self.r / 2),
                                     (self.x + self.r * 1.71, self.y), (self.x + self.r * 1.71, self.y + self.r),
                                     (self.x + self.r * 1.71 / 2, self.y + 3 * self.r / 2), (self.x, self.y + self.r)])

    def move_aim(self):
        """Moves the special aim in the screen (with reflection from the walls)"""
        self.color = random.choice(COLORS)
        self.x += 3 * self.speed_x / FPS
        self.y += 3 * self.speed_y / FPS
        self.r -= 1
        self.draw_aim()
        if self.r <= 10:
            self.color = random.choice(COLORS)
            self.x = randint(100, 1000)
            self.y = randint(100, 800)
            self.r = randint(50, 100)
            self.speed_x = randint(-200, 200)
            self.speed_y = randint(-200, 200)
        if self.x >= 1100:
            self.speed_x = randint(-100, -10)
        if self.x <= 50:
            self.speed_x = randint(10, 100)
        if self.y >= 800:
            self.speed_y = randint(-100, -10)
        if self.y <= 50:
            self.speed_y = randint(10, 100)

    def click_aim(self, pos):
        """Finds out if the player caught the aim"""
        x, y = pos
        if (self.x - x) ** 2 + (self.y - y) ** 2 <= self.r ** 2:
            self.color = random.choice(COLORS)
            self.x = randint(100, 1000)
            self.y = randint(100, 800)
            self.r = randint(50, 100)
            self.speed_x = randint(-200, 200)
            self.speed_y = randint(-200, 200)
            return True
        else:
            return False


class Car(object):

    def __init__(self):
        self.x = randint(200, 500)
        self.y = randint(200, 500)
        self.h = randint(10, 50)
        self.dir = 1
        self.speed_x = randint(10, 200)

    def draw_car(self):
        """Draws a car of a random color in random x, y coordinates with random size"""
        a = self.h / 50
        ellipse(screen, BLACK, (self.x - 15 * a, self.y + 35 * a, 30 * a, 10 * a))
        rect(screen, LIGHT_BLUE, (self.x, self.y, self.dir * 260 * a, self.h))
        rect(screen, LIGHT_BLUE, (self.x + self.dir * 40 * a, self.y - 40 * a, self.dir * 130 * a, 40 * a))
        rect(screen, LIGHT_GREY, (self.x + self.dir * 50 * a, self.y - 30 * a, self.dir * 45 * a, 30 * a))
        rect(screen, LIGHT_GREY, (self.x + self.dir * 120 * a, self.y - 30 * a, self.dir * 48 * a, 30 * a))
        rect(screen, LIGHT_GREY, (self.x + self.dir * 248 * a, self.y + 2 * a, self.dir * 10 * a, 10 * a))
        circle(screen, BLACK, (self.x + self.dir * int(220 * a), self.y + int(50 * a)), int(25 * a))
        circle(screen, BLACK, (self.x + self.dir * int(50 * a), self.y + int(50 * a)), int(25 * a))

    def move_car(self):
        """Moves the car in the screen (with reflection from the walls)"""
        a = self.h / 50
        self.x += self.speed_x / FPS
        if self.x + 170 * a >= 1100:
            self.dir = -1
            self.speed_x = -self.speed_x
        if self.x - 170 * a <= 50:
            self.dir = 1
            self.speed_x = -self.speed_x

    def click_car(self, pos):
        """Finds out if the player caught the car"""
        a = self.h / 50
        x, y = pos
        if ((x > self.x) and (x < self.x + 260 * a) and (y > self.y - 40 * a)
                and (y < self.y + self.h + 25 * a)):
            self.x = randint(200, 500)
            self.y = randint(200, 500)
            self.h = randint(10, 50)
            self.dir = 1
            self.speed_x = randint(10, 200)
            return True
        else:
            return False


def music():
    """Plays music for the game"""
    pygame.mixer.init()
    pygame.mixer.music.load("1.wav")
    pygame.mixer.music.play(100)


def write_leaders():
    """Writes leaderboards of top-10 players"""
    wb = openpyxl.load_workbook(filename='leaderboard.xlsx')
    sheet = wb['Sheet1']
    names = []
    ascores = []
    i = 1
    for j in range(10):
        names.append(str(sheet.cell(row=i, column=1).value))
        ascores.append(int(sheet.cell(row=i, column=2).value))
        i += 1
    names.append(your_name)
    ascores.append(int(str(scores.score)))
    leaders = dict(zip(ascores, names))
    lead_keys = list(leaders.keys())
    lead_keys = sorted(lead_keys, reverse=True)
    lead_items = []
    for k in lead_keys:
        lead_items.append((leaders[k]))
    lead_keys.pop()
    lead_items.pop()
    i = 1
    for p in lead_items:
        sheet.cell(row=i, column=1).value = p
        i += 1
    i = 1
    for m in lead_keys:
        sheet.cell(row=i, column=2).value = m
        i += 1
    wb.save('leaderboard.xlsx')


time = 0
game_length = 50

pygame.display.update()
clock = pygame.time.Clock()
finished = False

music()

intro = 0
while intro < 300:
    intro += 1
clock.tick(FPS)

scores = Score()
balls = [Ball() for i in range(10)]
aims = [Aim() for j in range(3)]
cars = [Car() for k in range(1)]

while time <= game_length * FPS and not finished:
    clock.tick(FPS)
    for ball in balls:
        ball.move()
        ball.draw_ball()
    for aim in aims:
        aim.move_aim()
        aim.draw_aim()
    for car in cars:
        car.draw_car()
        car.move_car()
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            finished = True
        elif event.type == pygame.MOUSEBUTTONDOWN:
            for ball in balls:
                if ball.click(event.pos):
                    scores.increase(1)
            for aim in aims:
                if aim.click_aim(event.pos):
                    scores.increase(int(1000 / aim.r))
            for car in cars:
                if car.click_car(event.pos):
                    scores.decrease()
    time += 1
    scores.text()
    scores.hello()
    pygame.display.update()
    screen.fill(WHITE)

pygame.quit()

your_name = input("Your name: ")
print("Your score: " + str(scores.score))

write_leaders()
